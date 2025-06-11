import os
import openpyxl
from pathlib import Path
import sys

# El email que quieres comprobar
EMAIL_A_VERIFICAR = "alb172@correo.ugr.es"  # Cambia esto a tu email

def diagnosticar_excel():
    print("\n===== DIAGNÓSTICO COMPLETO DE EXCEL =====")
    
    # 1. Buscar el archivo Excel
    print("\n1. BUSCANDO ARCHIVO EXCEL...")
    posibles_rutas = [
        Path(__file__).parent / "data" / "usuarios.xlsx",
        Path(__file__).parent / "usuarios.xlsx",
        Path(os.getcwd()) / "data" / "usuarios.xlsx",
        Path(os.getcwd()) / "usuarios.xlsx"
    ]
    
    excel_path = None
    for ruta in posibles_rutas:
        if ruta.exists():
            excel_path = ruta
            print(f"✅ Excel encontrado en: {excel_path}")
            break
    
    if not excel_path:
        print("❌ ERROR: No se encontró ningún archivo Excel")
        return
    
    # 2. Abrir el Excel directamente sin pandas
    print("\n2. ABRIENDO ARCHIVO EXCEL...")
    try:
        workbook = openpyxl.load_workbook(excel_path)
        print(f"✅ Excel cargado correctamente")
    except Exception as e:
        print(f"❌ ERROR al abrir Excel: {e}")
        return
    
    # 3. Examinar hojas
    print("\n3. HOJAS EN EL EXCEL:")
    sheets = workbook.sheetnames
    print(f"📑 Hojas encontradas: {sheets}")
    
    # 4. Examinar la primera hoja (asumimos que los datos están ahí)
    active_sheet = workbook[sheets[0]]
    print(f"\n4. EXAMINANDO HOJA: {active_sheet.title}")
    
    # 5. Buscar columna de email
    print("\n5. BUSCANDO COLUMNA DE EMAIL...")
    header_row = 1  # Asumimos que la primera fila tiene los encabezados
    email_col = None
    header_cells = []
    
    for col in range(1, active_sheet.max_column + 1):
        cell_value = active_sheet.cell(row=header_row, column=col).value
        header_cells.append(cell_value)
        if cell_value and 'mail' in str(cell_value).lower():
            email_col = col
            print(f"✅ Columna de email encontrada: '{cell_value}' en columna {col}")
    
    # Mostrar todos los encabezados para diagnóstico
    print(f"📋 Todos los encabezados: {header_cells}")
    
    if not email_col:
        print("❌ ERROR: No se encontró ninguna columna de email")
        return
    
    # 6. Buscar el email en la columna
    print(f"\n6. BUSCANDO EMAIL '{EMAIL_A_VERIFICAR}'...")
    email_encontrado = False
    todos_emails = []
    
    for row in range(2, active_sheet.max_row + 1):  # Empezamos en 2 para saltar encabezados
        cell_value = active_sheet.cell(row=row, column=email_col).value
        
        if cell_value:
            email_excel = str(cell_value).lower().strip()
            todos_emails.append(email_excel)
            
            if email_excel == EMAIL_A_VERIFICAR.lower().strip():
                email_encontrado = True
                print(f"✅ EMAIL ENCONTRADO en fila {row}")
    
    # 7. Mostrar resultados
    if not email_encontrado:
        print(f"❌ El email '{EMAIL_A_VERIFICAR}' NO ESTÁ en el Excel")
        print(f"📧 Emails encontrados en el Excel ({len(todos_emails)}):")
        for idx, email in enumerate(todos_emails, 1):
            print(f"  {idx}. '{email}'")
    else:
        print("\n✅ ¡ÉXITO! El email fue encontrado en el Excel")
    
    # 8. Sugerencias si no se encontró
    if not email_encontrado:
        print("\n🔍 SUGERENCIAS:")
        for email in todos_emails:
            if email in EMAIL_A_VERIFICAR or EMAIL_A_VERIFICAR in email:
                print(f"  - Posible coincidencia parcial: '{email}'")

if __name__ == "__main__":
    diagnosticar_excel()
    input("\nPresiona Enter para salir...")